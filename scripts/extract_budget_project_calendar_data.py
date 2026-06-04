import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\hanhc\Desktop\pm\项目信息.xlsx")
SOURCE_FALLBACK = Path(".tmp_project_info.xlsx")
GATEWAY_REFERENCE = Path("src/data/gatewayCalendar.json")
OUTPUT = Path("src/data/budgetProjectCalendar.json")

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
NODE_TYPES = ["BRD", "SRE", "GW1/2", "TUAT", "GW3/4/5"]
IMPORTANT_NODES = ["BRD", "SRE", "TUAT"]


def text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def iso_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value).strip()
    match = re.search(r"20\d{2}[-/]\d{1,2}[-/]\d{1,2}", raw)
    if match:
        year, month, day = re.split(r"[-/]", match.group(0))
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    return ""


def parse_iso_date(value):
    iso = iso_date(value)
    if re.match(r"20\d{2}-\d{2}-\d{2}$", iso):
        return datetime.strptime(iso, "%Y-%m-%d").date()
    return None


def normalize_name(value):
    return re.sub(r"\s+", " ", text(value)).strip().lower()


def has_valid_cscop(value):
    raw = text(value).strip()
    return bool(raw) and raw.upper() not in {"N/A", "NA", "NONE", "NULL", "-"}


def gateway_start_date(value):
    current = value + timedelta(days=1)
    while current.weekday() >= 5:
        current += timedelta(days=1)
    return current


def add_workdays(start, workdays):
    current = start
    remaining = workdays
    while remaining:
        if current.weekday() < 5:
            remaining -= 1
            if remaining == 0:
                break
        current += timedelta(days=1)
    return current


def month_for_node(start, end):
    if not start and not end:
        return ""
    start = start or end
    end = end or start
    if start.year == 2026:
        return MONTHS[start.month - 1]
    if start.year < 2026 <= end.year:
        return "Jan"
    return ""


def month_for_end_or_start(start, end):
    target = end or start
    if target and target.year == 2026:
        return MONTHS[target.month - 1]
    return ""


def detect_node(milestone, deliverable):
    milestone_raw = text(milestone).upper()
    deliverable_raw = text(deliverable).upper()
    if re.search(r"\bBRD\b", deliverable_raw):
        return "BRD"
    if re.search(r"\bSRE\b", deliverable_raw):
        return "SRE"
    if "DEVELOPMENT" in milestone_raw and re.search(r"\bTUAT\b|\bTUS\b|\bTUE\b", deliverable_raw):
        return "TUAT"
    return ""


def is_fsd_deliverable(deliverable):
    return bool(re.search(r"\bFSD\b", text(deliverable).upper()))


def row_dict(headers, row):
    return {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers)) if headers[idx]}


def project_key(row):
    return "|".join([
        text(row.get("所属Space")).lower(),
        text(row.get("项目编号")).lower(),
        normalize_name(row.get("项目名称")),
    ])


def load_gateway_reference():
    if not GATEWAY_REFERENCE.exists():
        return {}
    payload = json.loads(GATEWAY_REFERENCE.read_text(encoding="utf-8"))
    return {
        normalize_name(project.get("projectName")): project
        for project in payload.get("projects", [])
    }


def load_workbook():
    try:
        return openpyxl.load_workbook(SOURCE, read_only=True, data_only=True), SOURCE
    except (PermissionError, FileNotFoundError):
        return openpyxl.load_workbook(SOURCE_FALLBACK, read_only=True, data_only=True), SOURCE_FALLBACK


def main():
    wb, source_used = load_workbook()
    ws = wb["项目信息"]
    headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    gateway_reference = load_gateway_reference()
    projects = {}

    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row = row_dict(headers, raw_row)
        name = text(row.get("项目名称"))
        domain = text(row.get("所属Space"))
        if not name or not domain:
            continue

        key = project_key(row)
        reference_project = gateway_reference.get(normalize_name(name))
        if key not in projects:
            projects[key] = {
                "businessDomain": reference_project.get("businessDomain", domain) if reference_project else domain,
                "projectName": reference_project.get("projectName", name) if reference_project else name,
                "sourceProjectName": name,
                "cscopNo": reference_project.get("cscopNo", text(row.get("项目编号")) or "N/A") if reference_project else (text(row.get("项目编号")) or "N/A"),
                "sourceCscopNo": text(row.get("项目编号")) or "N/A",
                "status": reference_project.get("status", text(row.get("项目状态")) or "Unknown") if reference_project else (text(row.get("项目状态")) or "Unknown"),
                "contractOps": reference_project.get("contractOps", "") if reference_project else "",
                "handoverDone": reference_project.get("handoverDone", "") if reference_project else "",
                "gatewayReferenceMatched": bool(reference_project),
                "taskCount": 0,
                "taskStart": "",
                "taskEnd": "",
                "gatewayMonths": {month: [] for month in MONTHS},
                "tasks": [],
                "_nodes": {node_type: [] for node_type in IMPORTANT_NODES},
                "_sreEnds": [],
                "_fsdEnds": [],
                "_tuatEnds": [],
            }

        project = projects[key]
        start = iso_date(row.get("任务包开始时间"))
        end = iso_date(row.get("任务包结束时间"))
        task = {
            "milestone": text(row.get("阶段")) or "Unassigned",
            "deliverable": text(row.get("任务包")),
            "start": start,
            "end": end,
            "gw12Raw": "",
            "gw345Raw": "",
        }
        if any(task.values()):
            project["tasks"].append(task)
            project["taskCount"] += 1

        node_type = detect_node(row.get("阶段"), row.get("任务包"))
        task_start_date = parse_iso_date(row.get("任务包开始时间"))
        task_end_date = parse_iso_date(row.get("任务包结束时间"))
        if node_type in IMPORTANT_NODES:
            project["_nodes"][node_type].append({
                "start": task_start_date,
                "end": task_end_date,
                "milestone": task["milestone"],
                "deliverable": task["deliverable"] or node_type,
            })
            if node_type == "SRE" and task_start_date and task_end_date:
                project["_sreEnds"].append({"start": task_start_date, "end": task_end_date})
            if node_type == "TUAT" and task_start_date and task_end_date:
                project["_tuatEnds"].append({"start": task_start_date, "end": task_end_date})

        if is_fsd_deliverable(row.get("任务包")) and task_start_date and task_end_date:
            project["_fsdEnds"].append({"start": task_start_date, "end": task_end_date})

        dates = [item for item in [start, end, project["taskStart"], project["taskEnd"]] if item]
        project["taskStart"] = min(dates) if dates else ""
        project["taskEnd"] = max(dates) if dates else ""

    for project in projects.values():
        if not has_valid_cscop(project["cscopNo"]):
            continue

        for node_type in IMPORTANT_NODES:
            node_rows = [item for item in project["_nodes"][node_type] if item.get("start") or item.get("end")]
            if not node_rows:
                continue
            starts = [item["start"] for item in node_rows if item.get("start")]
            ends = [item["end"] for item in node_rows if item.get("end")]
            start = min(starts) if starts else None
            end = max(ends) if ends else start
            marker = {
                "type": node_type,
                "milestone": " / ".join(dict.fromkeys(item["milestone"] for item in node_rows if item["milestone"])) or node_type,
                "deliverable": node_type,
                "start": start.isoformat() if start else "",
                "end": end.isoformat() if end else "",
                "basis": "Merged project milestone from 项目信息",
            }
            month = month_for_node(start, end)
            if month:
                project["gatewayMonths"][month].append(marker)

        gw12_basis_rows = project["_sreEnds"] or project["_fsdEnds"]
        if gw12_basis_rows:
            ref = max(gw12_basis_rows, key=lambda item: item["end"])
            gw_start = gateway_start_date(ref["end"])
            gw_end = add_workdays(gw_start, 10)
            basis_type = "SRE" if project["_sreEnds"] else "FSD"
            marker = {
                "type": "GW1/2",
                "milestone": "Gateway Handover 1/2",
                "deliverable": "GW1/2 Handover",
                "start": gw_start.isoformat(),
                "end": gw_end.isoformat(),
                "basis": f"Calculated from {basis_type} end {ref['end'].isoformat()}",
            }
            month = month_for_end_or_start(gw_start, gw_end)
            if month:
                project["gatewayMonths"][month].append(marker)

        if project["_tuatEnds"]:
            ref = max(project["_tuatEnds"], key=lambda item: item["end"])
            gw_start = gateway_start_date(ref["end"])
            gw_end = add_workdays(gw_start, 5)
            marker = {
                "type": "GW3/4/5",
                "milestone": "Gateway Handover 3/4/5",
                "deliverable": "GW3/4/5 Handover",
                "start": gw_start.isoformat(),
                "end": gw_end.isoformat(),
                "basis": f"Calculated from TUAT end {ref['end'].isoformat()}",
            }
            month = month_for_end_or_start(gw_start, gw_end)
            if month:
                project["gatewayMonths"][month].append(marker)

    order = {"BRD": 1, "SRE": 2, "GW1/2": 3, "TUAT": 4, "GW3/4/5": 5}
    project_list = list(projects.values())
    for project in project_list:
        for key in ["_nodes", "_sreEnds", "_fsdEnds", "_tuatEnds"]:
            project.pop(key, None)
        for month in MONTHS:
            project["gatewayMonths"][month] = sorted(
                project["gatewayMonths"][month],
                key=lambda marker: (order.get(marker["type"], 99), marker.get("start") or "", marker.get("deliverable") or ""),
            )
        project["gatewayCount"] = sum(1 for items in project["gatewayMonths"].values() for marker in items if marker["type"] in ("GW1/2", "GW3/4/5"))
        project["nodeCount"] = sum(len(items) for items in project["gatewayMonths"].values())
        project["hasGateway"] = project["nodeCount"] > 0

    domain_counts = Counter(project["businessDomain"] for project in project_list)
    status_counts = Counter(project["status"] for project in project_list)
    month_counts = {month: {"projects": 0, **{node_type: 0 for node_type in NODE_TYPES}} for month in MONTHS}
    domain_month_counts = defaultdict(lambda: {month: {"projects": 0, **{node_type: 0 for node_type in NODE_TYPES}} for month in MONTHS})

    for project in project_list:
        for month, markers in project["gatewayMonths"].items():
            if markers:
                month_counts[month]["projects"] += 1
                domain_month_counts[project["businessDomain"]][month]["projects"] += 1
            for marker in markers:
                month_counts[month][marker["type"]] += 1
                domain_month_counts[project["businessDomain"]][month][marker["type"]] += 1

    summary = {
        "sourceFile": str(source_used),
        "projectCount": len(project_list),
        "taskCount": sum(project["taskCount"] for project in project_list),
        "projectsWithGateway": sum(1 for project in project_list if project["hasGateway"]),
        "projectsWithoutGateway": sum(1 for project in project_list if not project["hasGateway"]),
        "gatewayReferenceMatched": sum(1 for project in project_list if project.get("gatewayReferenceMatched")),
        "gatewayReferenceMissing": sum(1 for project in project_list if not project.get("gatewayReferenceMatched")),
        "domainCounts": dict(sorted(domain_counts.items())),
        "statusCounts": dict(sorted(status_counts.items())),
        "monthCounts": month_counts,
        "domainMonthCounts": {domain: counts for domain, counts in sorted(domain_month_counts.items())},
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({"summary": summary, "projects": project_list}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
