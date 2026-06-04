import json
from pathlib import Path

import openpyxl


SOURCE = Path(".tmp_wbs_template.xlsx")
OUTPUT = Path("src/data/wbsTemplate.json")


def text(value):
    if value is None:
        return ""
    return str(value).strip()


wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
ws = wb["WBS Template"]
headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=3, max_row=3))]
rows = []
current_stage = ""
current_epic = ""

for raw in ws.iter_rows(min_row=4, values_only=True):
    row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers)) if headers[i]}
    stage = text(row.get("Stage")) or current_stage
    epic = text(row.get("Epic")) or current_epic
    if text(row.get("Stage")):
        current_stage = stage
    if text(row.get("Epic")):
        current_epic = epic

    task = text(row.get("Task"))
    task_id = text(row.get("Task ID"))
    if not task_id and not task:
        continue

    rows.append({
        "stage": stage,
        "epic": epic,
        "taskId": task_id,
        "task": task,
        "scope": text(row.get("In Scope/Out of Scope")),
        "dependency": text(row.get("Dependency")),
        "gateway": text(row.get("Gateway checklist")),
        "owner": text(row.get("Owner（请把此列的角色更改为人名，每一个单元格只能有一个人）")),
        "supporter": text(row.get("Supporter（请在此列填写参与人）")),
        "status": text(row.get("Status")) or "Not Started",
        "manday": text(row.get("Manday")),
        "estimatedStart": text(row.get("Estimated Start Time")),
        "estimatedEnd": text(row.get("Estimated End Time")),
        "comment": text(row.get("Comment")),
        "riskAsk": text(row.get("Risk & Ask")),
        "jira": text(row.get("Jira Ticket （ Optional）")),
    })

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps({"items": rows}, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({
    "items": len(rows),
    "stages": sorted(set(item["stage"] for item in rows)),
}, ensure_ascii=False, indent=2))
