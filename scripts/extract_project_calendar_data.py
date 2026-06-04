import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\hanhc\Desktop\2026_Project_Calendar.xlsx")
OUTPUT = Path("src/data/projectCalendar.json")

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MILESTONE_ALIASES = {
    "BRD": "BRD",
    "WBS": "WBS",
    "SRE": "SRE",
    "SIT": "SIT",
    "TUAT": "TUAT",
    "TUS": "TUAT",
    "TUE": "TUAT",
    "BUA": "BUAT",
    "BUAT": "BUAT",
    "CAB": "CAB",
    "GATEWAY": "Gateway",
    "GW": "Gateway",
    "CUTOVER": "Cutover",
    "HANDOVER": "Handover",
    "HYPERCARE": "Hypercare",
    "HYPECARE": "Hypercare",
    "INTAKE": "Intake",
    "GOVERNANCE": "Governance",
    "DEPLOY": "Deploy",
    "CLOSURE": "Closure",
}


def text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def iso_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value).strip()
    if not raw:
        return ""
    match = re.search(r"20\d{2}[-/]\d{1,2}[-/]\d{1,2}", raw)
    if match:
        parts = re.split(r"[-/]", match.group(0))
        return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    return raw


def normalize_status(raw):
    clean = re.sub(r"[\u200e\u200f\u202a-\u202e\ufe0f]", "", text(raw))
    clean = re.sub(r"\s+", " ", clean).strip()
    for icon in ["🟢", "⚪", "🎉", "🔴", "🟡"]:
        clean = clean.replace(icon, "").strip()
    return clean or "Unknown"


def tokenize_milestones(raw):
    cleaned = re.sub(r"[\n\r,;/|]+", " ", text(raw))
    cleaned = cleaned.replace("🚀", " Handover ")
    found = []
    for token in re.findall(r"[A-Za-z][A-Za-z0-9-]*", cleaned):
        upper = token.upper()
        if upper.startswith("GW"):
            label = "Gateway"
        else:
            label = MILESTONE_ALIASES.get(upper)
        if label and label not in found:
            found.append(label)
    return found


def sheet_rows(ws, header_row):
    headers = [text(cell.value) for cell in ws[header_row]]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item = {}
        for idx, header in enumerate(headers):
            if header:
                item[header] = row[idx] if idx < len(row) else None
        yield item


def pick(row, *names):
    for name in names:
        if name in row and row.get(name) not in (None, ""):
            return row.get(name)
    return None


def project_key(name, cscop):
    return f"{text(name).lower()}::{text(cscop).lower()}"


def main():
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    full_ws = wb["2026 Full Calendar"]
    task_ws = wb["Timesheet Task Detail"]

    tasks_by_project = defaultdict(list)
    for row in sheet_rows(task_ws, 1):
        project_name = text(row.get("Full Calendar Project Name"))
        if not project_name:
            continue
        full_calendar_cscop = text(pick(row, "Full Calendar CSCOP No.", "Full Calendar CSCOP No")) or "N/A"
        task = {
            "businessDomain": text(row.get("Business Domain")),
            "projectName": project_name,
            "cscopNo": full_calendar_cscop,
            "issueKey": text(row.get("Issue Key")),
            "timesheetProjectName": text(row.get("Timesheet Project Name")),
            "timesheetCscopNo": text(row.get("Timesheet CSCOP No")),
            "timesheetStatus": text(row.get("????")),
            "milestone": text(row.get("Milestone")) or "Unassigned",
            "deliverable": text(row.get("Task Deliverables")),
            "calendarToken": text(row.get("Calendar Token")),
            "start": iso_date(row.get("Task Start")),
            "end": iso_date(row.get("Task End")),
            "months": [m.strip() for m in text(row.get("Calendar Months")).split(",") if m.strip()],
            "matchMethod": text(row.get("Match Method")),
        }
        tasks_by_project[project_key(project_name, full_calendar_cscop)].append(task)

    projects = []
    domain_counts = defaultdict(int)
    status_counts = defaultdict(int)
    milestone_counts = defaultdict(int)
    match_counts = defaultdict(int)

    for row in sheet_rows(full_ws, 2):
        project_name = text(row.get("Project Name"))
        business_domain = text(row.get("Business Domain"))
        if not project_name or not business_domain:
            continue

        cscop = text(row.get("CSCOP No.")) or "N/A"
        month_milestones = []
        all_milestones = []
        for month in MONTHS:
            raw = text(row.get(month))
            tokens = tokenize_milestones(raw)
            for token in tokens:
                if token not in all_milestones:
                    all_milestones.append(token)
                milestone_counts[token] += 1
            month_milestones.append({"month": month, "raw": raw, "tokens": tokens})

        tasks = tasks_by_project.get(project_key(project_name, cscop), [])
        task_dates = [d for t in tasks for d in [t["start"], t["end"]] if d]
        task_milestones = sorted({t["milestone"] for t in tasks if t["milestone"]})
        deliverables = sorted({t["deliverable"].strip() for t in tasks if t["deliverable"].strip()})
        timesheet_match = text(row.get("Timesheet Match"))

        if not all_milestones:
            all_milestones = tokenize_milestones(row.get("Milestones"))
        status = normalize_status(row.get("Status Trend") or row.get("Jira Status"))
        match_counts[timesheet_match or "No match"] += 1
        status_counts[status] += 1
        domain_counts[business_domain] += 1

        risks = []
        if not all_milestones:
            risks.append("No calendar milestone")
        if not tasks:
            risks.append("No task detail")
        if "handover" in " ".join([m["raw"] for m in month_milestones]).lower() and not any("handover" in t["milestone"].lower() or "handover" in t["deliverable"].lower() for t in tasks):
            risks.append("Handover needs task check")
        if "fuzzy" in timesheet_match.lower():
            risks.append("Fuzzy match")
        if cscop.upper() == "N/A":
            risks.append("Missing CSCOP")

        projects.append({
            "businessDomain": business_domain,
            "projectName": project_name,
            "cscopNo": cscop,
            "owner": text(row.get("Owner")),
            "issueKey": text(row.get("Issue Key")),
            "issueType": text(row.get("Issue Type")),
            "jiraStatus": text(row.get("Jira Status")) or "Unknown",
            "statusTrend": status,
            "startDate": iso_date(row.get("Start Date")),
            "goLiveDate": iso_date(row.get("Go-live Date")),
            "dueDate": iso_date(row.get("Due Date")),
            "parentKey": text(row.get("Parent Key")),
            "parentSummary": text(row.get("Parent Summary")),
            "calendarMatch": text(row.get("Calendar Match")),
            "matchScore": text(row.get("Match Score")),
            "matchSource": text(row.get("Match Source")),
            "sourceNotes": text(row.get("Source Notes")),
            "timesheetProjectStatus": text(row.get("Timesheet Project Status")),
            "timesheetTaskCount": len(tasks),
            "timesheetMatch": timesheet_match or "No match",
            "taskStart": min(task_dates) if task_dates else "",
            "taskEnd": max(task_dates) if task_dates else "",
            "calendarMilestones": all_milestones,
            "taskMilestones": task_milestones,
            "deliverables": deliverables,
            "monthMilestones": month_milestones,
            "tasks": tasks,
            "risks": risks,
        })

    handover_projects = sum(1 for p in projects if any("Handover" in m["tokens"] or "handover" in m["raw"].lower() for m in p["monthMilestones"]))
    summary = {
        "sourceFile": str(SOURCE),
        "projectCount": len(projects),
        "taskCount": sum(len(p["tasks"]) for p in projects),
        "withTaskDetail": sum(1 for p in projects if p["tasks"]),
        "withoutTaskDetail": sum(1 for p in projects if not p["tasks"]),
        "handoverProjects": handover_projects,
        "missingCscop": sum(1 for p in projects if p["cscopNo"].upper() == "N/A"),
        "noCalendarMilestone": sum(1 for p in projects if not p["calendarMilestones"]),
        "domains": dict(sorted(domain_counts.items())),
        "statuses": dict(sorted(status_counts.items())),
        "milestones": dict(sorted(milestone_counts.items(), key=lambda item: item[0])),
        "matchMethods": dict(sorted(match_counts.items())),
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({"summary": summary, "projects": projects}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
