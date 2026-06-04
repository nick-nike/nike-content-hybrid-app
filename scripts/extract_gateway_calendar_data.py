import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\hanhc\Desktop\Overall Projects Info - v1.1-260520.xlsx")
SOURCE_FALLBACK = Path(".tmp_overall_projects.xlsx")
CALENDAR_SOURCE = Path(r"C:\Users\hanhc\Desktop\2026_Project_Calendar.xlsx")
CALENDAR_FALLBACK = Path(".tmp_2026_Project_Calendar.xlsx")
OUTPUT = Path("src/data/gatewayCalendar.json")

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_ALIASES = {month.lower(): month for month in MONTHS}
IMPORTANT_NODES = ["BRD", "SRE", "TUAT"]
GATEWAY_NODES = {"GW1/2", "GW3/4/5"}


def text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def iso_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value).strip()
    match = re.search(r"20\d{2}[-/]\d{1,2}[-/]\d{1,2}", raw)
    if match:
        year, month, day = re.split(r"[-/]", match.group(0))
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    return raw


def parse_iso_date(value):
    iso = iso_date(value)
    if re.match(r"20\d{2}-\d{2}-\d{2}$", iso):
        return datetime.strptime(iso, "%Y-%m-%d").date()
    return None


def display_date(value):
    parsed = parse_iso_date(value)
    return parsed.isoformat() if parsed else ""


def date_from_month_day(value):
    raw = text(value)
    matches = re.findall(r"(?<!\d)(\d{1,2})[/-](\d{1,2})(?!\d)", raw)
    if not matches:
        return ""
    month, day = (int(part) for part in matches[-1])
    if 1 <= month <= 12 and 1 <= day <= 31:
        return f"2026-{month:02d}-{day:02d}"
    return ""


def gateway_start_date(value):
    current = value + timedelta(days=1)
    while current.weekday() >= 5:
        current += timedelta(days=1)
    return current


def add_workdays(start, workdays):
    current = start
    remaining = workdays
    while remaining:
        if current.weekday() < 5:
            remaining -= 1
            if remaining == 0:
                break
        current += timedelta(days=1)
    return current


def parse_months(value):
    raw = text(value)
    if not raw:
        return []
    found = []
    for token in re.findall(r"[A-Za-z]{3,9}|\d{1,2}", raw):
        key = token[:3].lower()
        if key in MONTH_ALIASES and MONTH_ALIASES[key] not in found:
            found.append(MONTH_ALIASES[key])
        elif token.isdigit():
            idx = int(token)
            if 1 <= idx <= 12 and MONTHS[idx - 1] not in found:
                found.append(MONTHS[idx - 1])
    return found


def months_from_dates(*values):
    found = []
    for value in values:
        iso = iso_date(value)
        if re.match(r"20\d{2}-\d{2}-\d{2}$", iso):
            month = MONTHS[int(iso[5:7]) - 1]
            if month not in found:
                found.append(month)
    return found


def months_between(start, end):
    if not start or not end:
        return []
    found = []
    current = date(start.year, start.month, 1)
    last = date(end.year, end.month, 1)
    while current <= last:
        if current.year == 2026:
            month = MONTHS[current.month - 1]
            if month not in found:
                found.append(month)
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    return found


def month_for_end_or_start(start, end):
    target = end or start
    if target and target.year == 2026:
        return MONTHS[target.month - 1]
    return ""


def month_for_node(start, end):
    if not start and not end:
        return ""
    start = start or end
    end = end or start
    if start.year == 2026:
        return MONTHS[start.month - 1]
    if start.year < 2026 < end.year:
        return "Jan"
    if start.year < 2026 and end.year == 2026:
        return "Jan"
    return ""


def detect_node(milestone, deliverable):
    milestone_raw = text(milestone).upper()
    deliverable_raw = text(deliverable).upper()
    if re.search(r"\bBRD\b", deliverable_raw):
        return "BRD"
    if re.search(r"\bSRE\b", deliverable_raw):
        return "SRE"
    if "DEVELOPMENT" in milestone_raw and re.search(r"\bTUAT\b|\bTUS\b|\bTUE\b", deliverable_raw):
        return "TUAT"
    return ""


def is_fsd_deliverable(deliverable):
    return bool(re.search(r"\bFSD\b", text(deliverable).upper()))


def row_dict(headers, row):
    return {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers)) if headers[idx]}


def get_value(row, *names):
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return value
    return None


def normalize_name(value):
    return re.sub(r"\s+", " ", text(value).replace("：", ":")).strip().lower()


def has_valid_cscop(value):
    raw = text(value).strip()
    return bool(raw) and raw.upper() not in {"N/A", "NA", "NONE", "NULL", "-"}


def load_workbook_with_fallback(primary, fallback):
    try:
        return openpyxl.load_workbook(primary, read_only=True, data_only=True), primary
    except PermissionError:
        return openpyxl.load_workbook(fallback, read_only=True, data_only=True), fallback


def load_calendar_supplements():
    try:
        wb, _source_used = load_workbook_with_fallback(CALENDAR_SOURCE, CALENDAR_FALLBACK)
    except FileNotFoundError:
        return {}
    if "2026 Full Calendar" not in wb.sheetnames:
        return {}

    ws = wb["2026 Full Calendar"]
    header_row = [text(cell.value) for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    supplements = {}
    month_indexes = {month: header_row.index(month) for month in MONTHS if month in header_row}
    for raw_row in ws.iter_rows(min_row=4, values_only=True):
        row = row_dict(header_row, raw_row)
        project_name = text(row.get("Project Name"))
        cscop_no = text(row.get("CSCOP No."))
        if not project_name or not cscop_no:
            continue

        start_date = display_date(row.get("Start Date"))
        go_live_date = display_date(row.get("Go-live Date"))
        due_date = display_date(row.get("Due Date"))
        planning = text(row.get("Planning"))
        candidates = []

        for month, idx in month_indexes.items():
            cell_text = text(raw_row[idx]).upper() if idx < len(raw_row) else ""
            if not cell_text:
                continue
            if re.search(r"\bBRD\b", cell_text):
                candidates.append(("BRD", month, start_date, go_live_date or due_date, "Full Calendar month cell"))
            if re.search(r"\bSRE\b", cell_text):
                candidates.append(("SRE", month, start_date, go_live_date or due_date, "Full Calendar month cell"))
            if re.search(r"\bTUAT\b|\bTUS\b|\bTUE\b", cell_text):
                candidates.append(("TUAT", month, start_date, go_live_date or due_date, "Full Calendar month cell"))
            if re.search(r"\bGW\s*1\s*/\s*2\b", cell_text):
                date_hint = date_from_month_day(cell_text)
                candidates.append(("GW1/2", month, date_hint or start_date, date_hint or go_live_date or due_date, "Full Calendar month cell"))
            if re.search(r"\bGW\s*3\s*/\s*4\s*/\s*5\b|\bHANDOVER\b", cell_text):
                candidates.append(("GW3/4/5", month, go_live_date or start_date, due_date or go_live_date, "Full Calendar month cell"))

        if re.search(r"\bGW\s*1\s*/\s*2\b", planning.upper()):
            date_hint = date_from_month_day(planning)
            target_month = MONTHS[int(date_hint[5:7]) - 1] if date_hint else ""
            if target_month:
                candidates.append(("GW1/2", target_month, date_hint, date_hint, "Full Calendar planning note"))
        if re.search(r"\bGW\s*3\s*/\s*4\s*/\s*5\b", planning.upper()):
            date_hint = date_from_month_day(planning)
            target_month = MONTHS[int(date_hint[5:7]) - 1] if date_hint else ""
            if target_month:
                candidates.append(("GW3/4/5", target_month, date_hint, date_hint, "Full Calendar planning note"))

        supplement = {
            "projectName": project_name,
            "cscopNo": cscop_no,
            "markers": [],
        }
        seen_types = set()
        for node_type, month, marker_start, marker_end, basis in candidates:
            if node_type in seen_types:
                continue
            if node_type in GATEWAY_NODES:
                continue
            seen_types.add(node_type)
            supplement["markers"].append({
                "month": month,
                "type": node_type,
                "milestone": basis,
                "deliverable": f"{node_type} Handover" if node_type.startswith("GW") else node_type,
                "start": marker_start,
                "end": marker_end,
                "basis": basis,
            })

        if supplement["markers"]:
            supplements[cscop_no.lower()] = supplement
            supplements[normalize_name(project_name)] = supplement
    return supplements


def project_key(row):
    return "|".join([
        text(row.get("Business Domain")).lower(),
        text(row.get("Project Name")).lower(),
        text(row.get("CSCOP No.")).lower(),
    ])


def main():
    source_used = SOURCE
    try:
        wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    except PermissionError:
        source_used = SOURCE_FALLBACK
        wb = openpyxl.load_workbook(SOURCE_FALLBACK, read_only=True, data_only=True)
    ws = wb["Project  Info"]
    headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    projects = {}

    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row = row_dict(headers, raw_row)
        name = text(row.get("Project Name"))
        domain = text(row.get("Business Domain"))
        if not name or not domain:
            continue

        key = project_key(row)
        if key not in projects:
            projects[key] = {
                "businessDomain": domain,
                "projectName": name,
                "cscopNo": text(row.get("CSCOP No.")) or "N/A",
                "status": text(row.get("项目状态")) or "Unknown",
                "contractOps": text(row.get("是否有合同（Ops）")),
                "handoverDone": text(row.get("是否已完成Handover")),
                "taskCount": 0,
                "taskStart": "",
                "taskEnd": "",
                "gatewayMonths": {month: [] for month in MONTHS},
                "tasks": [],
                "_nodes": {node_type: [] for node_type in IMPORTANT_NODES},
                "_sreEnds": [],
                "_fsdEnds": [],
                "_tuatEnds": [],
                "_manualGatewayRows": [],
            }

        project = projects[key]
        task = {
            "milestone": text(row.get("Milestone")) or "Unassigned",
            "deliverable": text(row.get("Task Deilverables")),
            "start": iso_date(row.get("Task Start")),
            "end": iso_date(row.get("Task End")),
            "gw12Raw": text(row.get("GW 1/2")),
            "gw345Raw": text(row.get("GW 3/4/5")),
        }
        if any(task.values()):
            project["tasks"].append(task)
            project["taskCount"] += 1

        node_type = detect_node(row.get("Milestone"), row.get("Task Deilverables"))
        if node_type in IMPORTANT_NODES:
            project["_nodes"][node_type].append({
                "start": parse_iso_date(row.get("Task Start")),
                "end": parse_iso_date(row.get("Task End")),
                "milestone": task["milestone"],
                "deliverable": task["deliverable"] or node_type,
            })
            task_start_date = parse_iso_date(row.get("Task Start"))
            task_end_date = parse_iso_date(row.get("Task End"))
            if node_type == "SRE" and task_start_date and task_end_date:
                project["_sreEnds"].append({
                    "start": task_start_date,
                    "end": task_end_date,
                    "milestone": task["milestone"],
                    "deliverable": task["deliverable"] or "SRE",
                })
            if node_type == "TUAT" and task_start_date and task_end_date:
                project["_tuatEnds"].append({
                    "start": task_start_date,
                    "end": task_end_date,
                    "milestone": task["milestone"],
                    "deliverable": task["deliverable"] or "TUAT",
                })

        task_start_date = parse_iso_date(row.get("Task Start"))
        task_end_date = parse_iso_date(row.get("Task End"))
        if is_fsd_deliverable(row.get("Task Deilverables")) and task_start_date and task_end_date:
            project["_fsdEnds"].append({
                "start": task_start_date,
                "end": task_end_date,
                "milestone": task["milestone"],
                "deliverable": task["deliverable"] or "FSD",
            })

        for gateway_type, column_name in [("GW1/2", "GW 1/2"), ("GW3/4/5", "GW 3/4/5")]:
            manual_months = parse_months(row.get(column_name))
            if manual_months:
                project["_manualGatewayRows"].append({
                    "type": gateway_type,
                    "months": manual_months,
                    "milestone": task["milestone"],
                    "deliverable": task["deliverable"],
                    "start": task["start"],
                    "end": task["end"],
                })

        dates = [task["start"], task["end"], project["taskStart"], project["taskEnd"]]
        dates = [item for item in dates if item]
        project["taskStart"] = min(dates) if dates else ""
        project["taskEnd"] = max(dates) if dates else ""

    calendar_supplements = load_calendar_supplements()
    for project in projects.values():
        for node_type in IMPORTANT_NODES:
            node_rows = [
                item for item in project["_nodes"][node_type]
                if item.get("start") or item.get("end")
            ]
            if node_rows:
                starts = [item["start"] for item in node_rows if item.get("start")]
                ends = [item["end"] for item in node_rows if item.get("end")]
                start = min(starts) if starts else None
                end = max(ends) if ends else start
                deliverables = []
                milestones = []
                for item in node_rows:
                    if item["deliverable"] and item["deliverable"] not in deliverables:
                        deliverables.append(item["deliverable"])
                    if item["milestone"] and item["milestone"] not in milestones:
                        milestones.append(item["milestone"])

                marker = {
                    "type": node_type,
                    "milestone": " / ".join(milestones) or node_type,
                    "deliverable": node_type,
                    "start": start.isoformat() if start else "",
                    "end": end.isoformat() if end else "",
                    "basis": "Merged project milestone",
                }
                month = month_for_node(start, end)
                if month and marker not in project["gatewayMonths"][month]:
                    project["gatewayMonths"][month].append(marker)

        gw12_basis_rows = project["_sreEnds"] or project["_fsdEnds"]
        if gw12_basis_rows:
            sre_ref = max(gw12_basis_rows, key=lambda item: item["end"])
            gw_start = gateway_start_date(sre_ref["end"])
            gw_end = add_workdays(gw_start, 10)
            basis_type = "SRE" if project["_sreEnds"] else "FSD"
            marker = {
                "type": "GW1/2",
                "milestone": "Gateway Handover 1/2",
                "deliverable": "GW1/2 Handover",
                "start": gw_start.isoformat(),
                "end": gw_end.isoformat(),
                "basis": f"Calculated from {basis_type} end {sre_ref['end'].isoformat()}",
            }
            month = month_for_end_or_start(gw_start, gw_end)
            if month and marker not in project["gatewayMonths"][month]:
                project["gatewayMonths"][month].append(marker)

        if project["_tuatEnds"]:
            tuat_ref = max(project["_tuatEnds"], key=lambda item: item["end"])
            gw_start = gateway_start_date(tuat_ref["end"])
            gw_end = add_workdays(gw_start, 5)
            marker = {
                "type": "GW3/4/5",
                "milestone": "Gateway Handover 3/4/5",
                "deliverable": "GW3/4/5 Handover",
                "start": gw_start.isoformat(),
                "end": gw_end.isoformat(),
                "basis": f"Calculated from TUAT end {tuat_ref['end'].isoformat()}",
            }
            month = month_for_end_or_start(gw_start, gw_end)
            if month and marker not in project["gatewayMonths"][month]:
                project["gatewayMonths"][month].append(marker)

        supplement = calendar_supplements.get(project["cscopNo"].lower()) if has_valid_cscop(project["cscopNo"]) else None
        if supplement:
            existing_types = {
                marker["type"]
                for month in MONTHS
                for marker in project["gatewayMonths"][month]
            }
            for supplement_marker in supplement["markers"]:
                if supplement_marker["type"] in existing_types:
                    continue
                month = supplement_marker["month"]
                marker = {key: value for key, value in supplement_marker.items() if key != "month"}
                if month in project["gatewayMonths"] and marker not in project["gatewayMonths"][month]:
                    project["gatewayMonths"][month].append(marker)
                    existing_types.add(marker["type"])
                    for date_value in (marker.get("start"), marker.get("end")):
                        if date_value:
                            dates = [date_value, project["taskStart"], project["taskEnd"]]
                            dates = [item for item in dates if item]
                            project["taskStart"] = min(dates) if dates else ""
                            project["taskEnd"] = max(dates) if dates else ""

    order = {"BRD": 1, "SRE": 2, "GW1/2": 3, "TUAT": 4, "GW3/4/5": 5}
    project_list = list(projects.values())
    for project in project_list:
        project.pop("_nodes", None)
        project.pop("_sreEnds", None)
        project.pop("_fsdEnds", None)
        project.pop("_tuatEnds", None)
        project.pop("_manualGatewayRows", None)
        for month in MONTHS:
            project["gatewayMonths"][month] = sorted(
                project["gatewayMonths"][month],
                key=lambda marker: (order.get(marker["type"], 99), marker.get("start") or "", marker.get("deliverable") or ""),
            )
        project["gatewayCount"] = sum(1 for items in project["gatewayMonths"].values() for marker in items if marker["type"] in ("GW1/2", "GW3/4/5"))
        project["nodeCount"] = sum(len(items) for items in project["gatewayMonths"].values())
        project["hasGateway"] = project["nodeCount"] > 0

    domain_counts = Counter(project["businessDomain"] for project in project_list)
    status_counts = Counter(project["status"] for project in project_list)
    node_types = ["BRD", "SRE", "GW1/2", "TUAT", "GW3/4/5"]
    month_counts = {month: {"projects": 0, **{node_type: 0 for node_type in node_types}} for month in MONTHS}
    domain_month_counts = defaultdict(lambda: {month: {"projects": 0, **{node_type: 0 for node_type in node_types}} for month in MONTHS})

    for project in project_list:
        for month, markers in project["gatewayMonths"].items():
            if markers:
                month_counts[month]["projects"] += 1
                domain_month_counts[project["businessDomain"]][month]["projects"] += 1
            for marker in markers:
                month_counts[month][marker["type"]] += 1
                domain_month_counts[project["businessDomain"]][month][marker["type"]] += 1

    summary = {
        "sourceFile": str(source_used),
        "projectCount": len(project_list),
        "taskCount": sum(project["taskCount"] for project in project_list),
        "projectsWithGateway": sum(1 for project in project_list if project["hasGateway"]),
        "projectsWithoutGateway": sum(1 for project in project_list if not project["hasGateway"]),
        "domainCounts": dict(sorted(domain_counts.items())),
        "statusCounts": dict(sorted(status_counts.items())),
        "monthCounts": month_counts,
        "domainMonthCounts": {domain: counts for domain, counts in sorted(domain_month_counts.items())},
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({"summary": summary, "projects": project_list}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
